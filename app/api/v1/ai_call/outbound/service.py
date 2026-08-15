from __future__ import annotations

import asyncio
import json
import os
import re
import tempfile
import zipfile
from collections import Counter
from datetime import datetime, timezone
from itertools import islice
from pathlib import Path
from typing import Any

import aiofiles
from fastapi import UploadFile, status
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy import Integer, and_, delete, func, or_, select
from sqlalchemy.exc import SQLAlchemyError
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker

from app.api.v1.ai_call.model import AiCallPromptProfileModel
from app.core.exceptions import CustomException
from app.core.logger import log
from app.utils.id_util import generate_snowflake_id

from .model import AiCallOutboundValidationModel, AiCallOutboundValidationRowModel
from .schema import BatchValidationRequest, ValidationIssueOut, ValidationResultOut
from .sip_line_service import SipLineService

MAX_UPLOAD_BYTES = 10 * 1024 * 1024
UPLOAD_CHUNK_BYTES = 1024 * 1024
DEFAULT_PARSE_BATCH_SIZE = 500
TEMP_FILE_PREFIX = "ai-call-outbound-"
PHONE_PATTERN = re.compile(r"^1[3-9]\d{9}$")
PROMPT_VARIABLE_RE = re.compile(r"\{\{([A-Za-z][A-Za-z0-9_]*)\}\}")
LEGACY_VARIABLE_LABELS = {"customerName": "客户名称"}


def _now() -> datetime:
    return datetime.now(timezone.utc)


class OutboundValidationService:
    """批量名单上传、流式解析和租户隔离查询。"""

    def __init__(
        self,
        session_factory: async_sessionmaker[AsyncSession],
        *,
        parse_batch_size: int = DEFAULT_PARSE_BATCH_SIZE,
        line_service: SipLineService | None = None,
    ) -> None:
        self.session_factory = session_factory
        self.parse_batch_size = max(1, parse_batch_size)
        self.line_service = line_service or SipLineService()
        self._tasks: set[asyncio.Task] = set()

    async def accept_batch(
        self,
        *,
        db: AsyncSession,
        tenant_id: str,
        user_id: int,
        file: UploadFile,
        request: BatchValidationRequest,
    ) -> AiCallOutboundValidationModel:
        filename = (file.filename or "").strip()
        if Path(filename).suffix.lower() != ".xlsx":
            raise CustomException(
                msg="只支持单个 .xlsx 名单文件",
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        line = await self.line_service.resolve_default(db, tenant_id)
        temp_path: str | None = None
        try:
            temp_path, file_size = await self._save_upload(file)
            now = _now()
            validation = AiCallOutboundValidationModel(
                id=generate_snowflake_id(),
                tenant_id=tenant_id,
                status="VALIDATING",
                processing_stage="UPLOADED",
                original_filename=filename,
                temp_file_path=temp_path,
                file_size=file_size,
                task_config_json=json.dumps(
                    request.model_dump(mode="json", by_alias=True),
                    ensure_ascii=False,
                    separators=(",", ":"),
                ),
                line_id=line.id,
                line_snapshot_json=self.line_service.snapshot_json(line),
                valid_target_count=0,
                issue_count=0,
                issue_stats_json="{}",
                error_message=None,
                retryable=False,
                retry_count=0,
                created_by=user_id,
                created_at=now,
                updated_at=now,
                finished_at=None,
            )
            db.add(validation)
            await db.flush()
            return validation
        except Exception:
            if temp_path:
                self._delete_temp_file(temp_path)
            raise

    async def _save_upload(self, file: UploadFile) -> tuple[str, int]:
        descriptor, temp_path = tempfile.mkstemp(prefix=TEMP_FILE_PREFIX, suffix=".xlsx")
        os.close(descriptor)
        total = 0
        try:
            async with aiofiles.open(temp_path, "wb") as output:
                while chunk := await file.read(UPLOAD_CHUNK_BYTES):
                    total += len(chunk)
                    if total > MAX_UPLOAD_BYTES:
                        raise CustomException(
                            msg="名单文件不能超过 10 MB",
                            status_code=status.HTTP_413_CONTENT_TOO_LARGE,
                        )
                    await output.write(chunk)
            if total == 0:
                raise CustomException(
                    msg="名单文件不能为空",
                    status_code=status.HTTP_400_BAD_REQUEST,
                )
            return temp_path, total
        except Exception:
            self._delete_temp_file(temp_path)
            raise

    def schedule_validation(self, tenant_id: str, validation_id: int) -> None:
        task = asyncio.create_task(self.process_validation(tenant_id, validation_id))
        self._tasks.add(task)
        task.add_done_callback(self._tasks.discard)

    async def process_validation(self, tenant_id: str, validation_id: int) -> None:
        async with self.session_factory() as db:
            validation = await self._get_validation(db, tenant_id, validation_id)
            if validation.status != "VALIDATING":
                return

            if validation.processing_stage == "PARSED":
                if self._delete_temp_file(validation.temp_file_path):
                    validation.temp_file_path = None
                await db.commit()
                await self._finish_from_persisted_rows(db, validation)
                return

            temp_path = validation.temp_file_path
            if not temp_path or not Path(temp_path).is_file():
                await self._mark_parse_failure(
                    db,
                    validation,
                    "临时名单文件不存在，请重新上传完整名单",
                )
                return

            validation.processing_stage = "PARSING"
            validation.updated_at = _now()
            await db.execute(
                delete(AiCallOutboundValidationRowModel).where(
                    AiCallOutboundValidationRowModel.tenant_id == tenant_id,
                    AiCallOutboundValidationRowModel.validation_id == validation_id,
                )
            )
            await db.commit()

            try:
                await self._parse_rows(db, validation, temp_path)
            except (InvalidFileException, OSError, ValueError, KeyError, zipfile.BadZipFile) as exc:
                await db.rollback()
                validation = await self._get_validation(db, tenant_id, validation_id)
                await self._mark_parse_failure(
                    db,
                    validation,
                    f"名单解析失败，请重新上传完整名单：{type(exc).__name__}",
                )
                return
            except Exception as exc:
                await db.rollback()
                validation = await self._get_validation(db, tenant_id, validation_id)
                await self._mark_parse_failure(
                    db,
                    validation,
                    f"名单解析失败，请重新上传完整名单：{type(exc).__name__}",
                )
                log.exception(f"外呼名单解析失败 validation_id={validation_id}: {exc}")
                return

            validation.processing_stage = "PARSED"
            validation.updated_at = _now()
            await db.commit()
            if self._delete_temp_file(temp_path):
                validation.temp_file_path = None
            await db.commit()
            await self._finish_from_persisted_rows(db, validation)

    async def _parse_rows(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel,
        temp_path: str,
    ) -> None:
        workbook = await asyncio.to_thread(
            load_workbook,
            temp_path,
            read_only=True,
            data_only=True,
        )
        try:
            sheet = workbook.active
            rows = iter(sheet.iter_rows(values_only=True))
            header_values = await asyncio.to_thread(lambda: next(rows, None))
            if header_values is None:
                await self._persist_batch(
                    db,
                    validation,
                    [self._issue_row(1, None, None, ["名单文件没有表头"])],
                )
                return

            variable_columns = await self._prompt_variable_columns(db, validation)
            header_map, header_reasons = self._header_map(
                header_values,
                variable_columns,
            )
            if header_reasons:
                await self._persist_batch(
                    db,
                    validation,
                    [self._issue_row(1, None, None, header_reasons)],
                )
                return

            found_data_row = False
            row_number = 2
            while True:
                batch_values = await asyncio.to_thread(
                    lambda: list(islice(rows, self.parse_batch_size))
                )
                if not batch_values:
                    break
                batch = [
                    self._build_row(
                        row_number + offset,
                        values,
                        header_map,
                        variable_columns,
                    )
                    for offset, values in enumerate(batch_values)
                    if not self._is_empty_row(values)
                ]
                if batch:
                    found_data_row = True
                    await self._persist_batch(db, validation, batch)
                row_number += len(batch_values)

            if not found_data_row:
                await self._persist_batch(
                    db,
                    validation,
                    [self._issue_row(2, None, None, ["名单不包含数据行"])],
                )
        finally:
            await asyncio.to_thread(workbook.close)

    @staticmethod
    def _header_map(
        values: tuple[Any, ...],
        variable_columns: dict[str, str],
    ) -> tuple[dict[str, int], list[str]]:
        normalized = [OutboundValidationService._cell_text(value) for value in values]
        supported_headers = {"手机号", "客户名称", *variable_columns}
        header_map = {
            header: index
            for index, header in enumerate(normalized)
            if header in supported_headers
        }
        reasons: list[str] = []
        if "手机号" not in header_map:
            reasons.append("缺少必填表头：手机号")
        unsupported = [
            header for header in normalized if header and header not in supported_headers
        ]
        if unsupported:
            reasons.append(f"存在不支持的表头：{'、'.join(unsupported)}")
        if len([header for header in normalized if header == "手机号"]) > 1:
            reasons.append("手机号表头重复")
        for label in variable_columns:
            if label not in header_map:
                reasons.append(f"缺少必填表头：{label}")
            if normalized.count(label) > 1:
                reasons.append(f"{label}表头重复")
        return header_map, reasons

    def _build_row(
        self,
        row_number: int,
        values: tuple[Any, ...],
        header_map: dict[str, int],
        variable_columns: dict[str, str],
    ) -> dict[str, Any]:
        phone = self._value_at(values, header_map["手机号"])
        business_params = {
            key: self._value_at(values, header_map.get(label))
            for label, key in variable_columns.items()
        }
        customer_name = business_params.get("customerName") or self._value_at(
            values,
            header_map.get("客户名称"),
        )
        if customer_name and "customerName" not in business_params:
            business_params["customerName"] = customer_name
        reasons: list[str] = []
        if self._is_empty_row(values):
            reasons.append("空行")
        elif not phone:
            reasons.append("手机号不能为空")
        elif not PHONE_PATTERN.fullmatch(phone):
            reasons.append("手机号格式错误")
        if customer_name and len(customer_name) > 100:
            reasons.append("客户名称不能超过100个字符")
        for label, key in variable_columns.items():
            if not business_params.get(key):
                reasons.append(f"{label}不能为空")
        return {
            "id": generate_snowflake_id(),
            "tenant_id": "",
            "validation_id": 0,
            "row_number": row_number,
            "phone_number": phone or None,
            "customer_name": customer_name or None,
            "business_params_json": json.dumps(
                business_params,
                ensure_ascii=False,
                separators=(",", ":"),
            ),
            "normalized_phone": phone if PHONE_PATTERN.fullmatch(phone) else None,
            "is_valid": not reasons,
            "reasons": reasons,
            "duplicate_row_number": None,
            "created_at": _now(),
        }

    @classmethod
    def _is_empty_row(cls, values: tuple[Any, ...]) -> bool:
        return not any(cls._cell_text(value) for value in values)

    @staticmethod
    def _issue_row(
        row_number: int,
        phone_number: str | None,
        customer_name: str | None,
        reasons: list[str],
    ) -> dict[str, Any]:
        return {
            "id": generate_snowflake_id(),
            "tenant_id": "",
            "validation_id": 0,
            "row_number": row_number,
            "phone_number": phone_number,
            "customer_name": customer_name,
            "business_params_json": "{}",
            "normalized_phone": None,
            "is_valid": False,
            "reasons": reasons,
            "duplicate_row_number": None,
            "created_at": _now(),
        }

    async def _persist_batch(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel,
        batch: list[dict[str, Any]],
    ) -> None:
        phone_groups: dict[str, list[dict[str, Any]]] = {}
        for row in batch:
            normalized_phone = row["normalized_phone"]
            if normalized_phone:
                phone_groups.setdefault(normalized_phone, []).append(row)

        first_existing_by_phone: dict[str, AiCallOutboundValidationRowModel] = {}
        if phone_groups:
            first_row_per_phone = (
                select(
                    AiCallOutboundValidationRowModel.normalized_phone.label("phone"),
                    func.min(AiCallOutboundValidationRowModel.row_number).label("row_number"),
                )
                .where(
                    AiCallOutboundValidationRowModel.tenant_id == validation.tenant_id,
                    AiCallOutboundValidationRowModel.validation_id == validation.id,
                    AiCallOutboundValidationRowModel.normalized_phone.in_(phone_groups),
                )
                .group_by(AiCallOutboundValidationRowModel.normalized_phone)
                .subquery()
            )
            first_existing_rows = (
                await db.scalars(
                    select(AiCallOutboundValidationRowModel)
                    .join(
                        first_row_per_phone,
                        and_(
                            AiCallOutboundValidationRowModel.normalized_phone
                            == first_row_per_phone.c.phone,
                            AiCallOutboundValidationRowModel.row_number
                            == first_row_per_phone.c.row_number,
                        ),
                    )
                    .order_by(AiCallOutboundValidationRowModel.row_number)
                )
            ).all()
            first_existing_by_phone = {
                row.normalized_phone or "": row for row in first_existing_rows
            }

        for phone, new_rows in phone_groups.items():
            first_existing = first_existing_by_phone.get(phone)
            if first_existing is not None:
                first_new = new_rows[0]
                reasons = self._reasons(first_existing.reasons_json)
                if "手机号重复" not in reasons:
                    reasons.append("手机号重复")
                first_existing.reasons_json = self._reasons_json(reasons)
                first_existing.is_valid = False
                first_existing.duplicate_row_number = (
                    first_existing.duplicate_row_number or first_new["row_number"]
                )
                for new_row in new_rows:
                    self._add_duplicate_reason(new_row, first_existing.row_number)
            elif len(new_rows) > 1:
                for index, new_row in enumerate(new_rows):
                    related = new_rows[1 if index == 0 else 0]["row_number"]
                    self._add_duplicate_reason(new_row, related)

        for row in batch:
            row["tenant_id"] = validation.tenant_id
            row["validation_id"] = validation.id
            reasons = row.pop("reasons")
            row["reasons_json"] = self._reasons_json(reasons) if reasons else None
            db.add(AiCallOutboundValidationRowModel(**row))
        await db.commit()

    @staticmethod
    def _add_duplicate_reason(row: dict[str, Any], related_row_number: int) -> None:
        if "手机号重复" not in row["reasons"]:
            row["reasons"].append("手机号重复")
        row["is_valid"] = False
        row["duplicate_row_number"] = related_row_number

    async def _finish_from_persisted_rows(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel,
    ) -> None:
        try:
            await self._finalize_validation(db, validation)
            await db.commit()
        except Exception as exc:
            await db.rollback()
            validation = await self._get_validation(db, validation.tenant_id, validation.id)
            validation.status = "SYSTEM_ERROR"
            validation.processing_stage = "PARSED"
            validation.error_message = "名单已解析，但系统校验失败，可使用 validationId 重试"
            validation.retryable = True
            validation.updated_at = _now()
            validation.finished_at = _now()
            await db.commit()
            log.exception(f"外呼名单系统校验失败 validation_id={validation.id}: {exc}")

    async def _finalize_validation(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel,
    ) -> None:
        counts = await db.execute(
            select(
                func.count(AiCallOutboundValidationRowModel.id),
                func.sum(func.cast(AiCallOutboundValidationRowModel.is_valid, Integer)),
            ).where(
                AiCallOutboundValidationRowModel.tenant_id == validation.tenant_id,
                AiCallOutboundValidationRowModel.validation_id == validation.id,
            )
        )
        total_count, valid_count = counts.one()
        valid_count = int(valid_count or 0)
        issue_count = int(total_count or 0) - valid_count
        issue_stats: Counter[str] = Counter()
        stream = await db.stream_scalars(
            select(AiCallOutboundValidationRowModel.reasons_json).where(
                AiCallOutboundValidationRowModel.tenant_id == validation.tenant_id,
                AiCallOutboundValidationRowModel.validation_id == validation.id,
                AiCallOutboundValidationRowModel.is_valid.is_(False),
            )
        )
        async for reasons_json in stream:
            issue_stats.update(self._reasons(reasons_json))

        validation.valid_target_count = valid_count
        validation.issue_count = issue_count
        validation.issue_stats_json = json.dumps(
            dict(issue_stats),
            ensure_ascii=False,
            separators=(",", ":"),
        )
        validation.status = "FAILED" if issue_count else "PASSED"
        validation.processing_stage = "COMPLETED"
        validation.error_message = None
        validation.retryable = False
        validation.updated_at = _now()
        validation.finished_at = _now()

    async def _mark_parse_failure(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel,
        message: str,
    ) -> None:
        if self._delete_temp_file(validation.temp_file_path):
            validation.temp_file_path = None
        validation.status = "SYSTEM_ERROR"
        validation.processing_stage = "PARSE_FAILED"
        validation.error_message = message
        validation.retryable = False
        validation.updated_at = _now()
        validation.finished_at = _now()
        await db.commit()

    async def prepare_retry(
        self,
        db: AsyncSession,
        tenant_id: str,
        validation_id: int,
    ) -> AiCallOutboundValidationModel:
        validation = await self._get_validation(db, tenant_id, validation_id)
        if (
            validation.status != "SYSTEM_ERROR"
            or validation.processing_stage != "PARSED"
            or not validation.retryable
        ):
            raise CustomException(
                msg="当前校验不能重试，请重新上传完整名单",
                status_code=status.HTTP_409_CONFLICT,
            )
        validation.status = "VALIDATING"
        validation.error_message = None
        validation.retryable = False
        validation.retry_count += 1
        validation.updated_at = _now()
        validation.finished_at = None
        await db.flush()
        return validation

    async def get_result(
        self,
        db: AsyncSession,
        tenant_id: str,
        validation_id: int,
    ) -> ValidationResultOut:
        validation = await self._get_validation(db, tenant_id, validation_id)
        return self.result_out(validation)

    async def list_issues(
        self,
        db: AsyncSession,
        tenant_id: str,
        validation_id: int,
        *,
        page_num: int,
        page_size: int,
        phone_number: str | None,
        reason: str | None,
    ) -> tuple[list[ValidationIssueOut], int]:
        await self._get_validation(db, tenant_id, validation_id)
        conditions = [
            AiCallOutboundValidationRowModel.tenant_id == tenant_id,
            AiCallOutboundValidationRowModel.validation_id == validation_id,
            AiCallOutboundValidationRowModel.is_valid.is_(False),
        ]
        if phone_number:
            conditions.append(
                AiCallOutboundValidationRowModel.phone_number.contains(phone_number.strip())
            )
        if reason:
            conditions.append(
                AiCallOutboundValidationRowModel.reasons_json.contains(reason.strip())
            )
        total = int(
            (
                await db.scalar(
                    select(func.count(AiCallOutboundValidationRowModel.id)).where(*conditions)
                )
            )
            or 0
        )
        rows = (
            await db.scalars(
                select(AiCallOutboundValidationRowModel)
                .where(*conditions)
                .order_by(AiCallOutboundValidationRowModel.row_number)
                .offset((page_num - 1) * page_size)
                .limit(page_size)
            )
        ).all()
        return [self.issue_out(row) for row in rows], total

    async def build_issue_export(
        self,
        db: AsyncSession,
        tenant_id: str,
        validation_id: int,
        *,
        phone_number: str | None,
        reason: str | None,
    ) -> str:
        await self._get_validation(db, tenant_id, validation_id)
        conditions = [
            AiCallOutboundValidationRowModel.tenant_id == tenant_id,
            AiCallOutboundValidationRowModel.validation_id == validation_id,
            AiCallOutboundValidationRowModel.is_valid.is_(False),
        ]
        if phone_number:
            conditions.append(
                AiCallOutboundValidationRowModel.phone_number.contains(phone_number.strip())
            )
        if reason:
            conditions.append(
                AiCallOutboundValidationRowModel.reasons_json.contains(reason.strip())
            )
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("问题明细")
        sheet.append(["原文件行号", "手机号", "客户名称", "错误原因", "重复位置"])
        stream = await db.stream_scalars(
            select(AiCallOutboundValidationRowModel)
            .where(*conditions)
            .order_by(AiCallOutboundValidationRowModel.row_number)
            .execution_options(yield_per=self.parse_batch_size)
        )
        async for row in stream:
            sheet.append([
                row.row_number,
                row.phone_number or "",
                row.customer_name or "",
                "；".join(self._reasons(row.reasons_json)),
                str(row.duplicate_row_number or ""),
            ])
        descriptor, export_path = tempfile.mkstemp(
            prefix="ai-call-outbound-issues-",
            suffix=".xlsx",
        )
        os.close(descriptor)
        try:
            await asyncio.to_thread(workbook.save, export_path)
        except Exception:
            self._delete_temp_file(export_path)
            raise
        return export_path

    async def recover_pending(self) -> None:
        try:
            async with self.session_factory() as db:
                pending = (
                    await db.execute(
                        select(
                            AiCallOutboundValidationModel.tenant_id,
                            AiCallOutboundValidationModel.id,
                            AiCallOutboundValidationModel.status,
                            AiCallOutboundValidationModel.processing_stage,
                            AiCallOutboundValidationModel.temp_file_path,
                        ).where(
                            or_(
                                AiCallOutboundValidationModel.status == "VALIDATING",
                                AiCallOutboundValidationModel.temp_file_path.is_not(None),
                            )
                        )
                    )
                ).all()
                for (
                    tenant_id,
                    validation_id,
                    validation_status,
                    processing_stage,
                    temp_file_path,
                ) in pending:
                    if validation_status != "VALIDATING":
                        if self._delete_temp_file(temp_file_path):
                            validation = await self._get_validation(
                                db,
                                tenant_id,
                                validation_id,
                            )
                            validation.temp_file_path = None
                            validation.updated_at = _now()
                            await db.commit()
                        continue
                    if processing_stage == "PARSED":
                        self.schedule_validation(tenant_id, validation_id)
                        continue
                    if temp_file_path and Path(temp_file_path).is_file():
                        self.schedule_validation(tenant_id, validation_id)
                        continue
                    validation = await self._get_validation(db, tenant_id, validation_id)
                    await self._mark_parse_failure(
                        db,
                        validation,
                        "服务重启后临时名单文件不存在，请重新上传完整名单",
                    )
        except (SQLAlchemyError, OSError) as exc:
            log.warning(f"跳过通用外呼名单恢复扫描，数据表尚未就绪: {type(exc).__name__}")

    async def create_template(
        self,
        db: AsyncSession,
        tenant_id: str,
        prompt_profile_id: int | None = None,
    ) -> str:
        labels = ["客户名称"]
        if prompt_profile_id is not None:
            profile = await db.scalar(
                select(AiCallPromptProfileModel).where(
                    AiCallPromptProfileModel.tenant_id == tenant_id,
                    AiCallPromptProfileModel.id == prompt_profile_id,
                )
            )
            if profile is None:
                raise CustomException(msg="提示词配置不存在", status_code=404)
            labels = list(
                (await self._prompt_variable_columns(db, profile=profile)).keys()
            )
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet("外呼名单")
        sheet.append(["手机号", *labels])
        descriptor, template_path = tempfile.mkstemp(
            prefix="ai-call-outbound-template-",
            suffix=".xlsx",
        )
        os.close(descriptor)
        try:
            workbook.save(template_path)
        except Exception:
            OutboundValidationService._delete_temp_file(template_path)
            raise
        return template_path

    async def _prompt_variable_columns(
        self,
        db: AsyncSession,
        validation: AiCallOutboundValidationModel | None = None,
        *,
        profile: AiCallPromptProfileModel | None = None,
    ) -> dict[str, str]:
        if profile is None:
            config = self._load_config(validation.task_config_json if validation else "{}")
            prompt_profile_id = config.get("promptProfileId")
            conditions = [
                AiCallPromptProfileModel.tenant_id == validation.tenant_id,
                AiCallPromptProfileModel.scene_code == config.get("sceneCode"),
            ]
            if prompt_profile_id:
                conditions.append(AiCallPromptProfileModel.id == int(prompt_profile_id))
            profile = await db.scalar(select(AiCallPromptProfileModel).where(*conditions))
        if profile is None:
            # 兼容尚未配置提示词的旧任务；新任务在创建阶段仍会校验所选配置。
            return {}
        try:
            variables = json.loads(profile.variables_json or "[]")
        except json.JSONDecodeError:
            variables = []
        labels_by_key = {
            str(item.get("key")): str(item.get("label"))
            for item in variables
            if isinstance(item, dict) and item.get("key") and item.get("label")
        }
        referenced = set(
            PROMPT_VARIABLE_RE.findall(
                "\n".join(
                    [
                        profile.opening_message or "",
                        profile.product_info or "",
                        profile.prompt_text or "",
                    ]
                )
            )
        )
        return {
            labels_by_key.get(key) or LEGACY_VARIABLE_LABELS.get(key) or key: key
            for key in sorted(referenced)
        }

    @staticmethod
    def _load_config(value: str) -> dict[str, Any]:
        try:
            result = json.loads(value)
        except json.JSONDecodeError:
            return {}
        return result if isinstance(result, dict) else {}

    @staticmethod
    def result_out(validation: AiCallOutboundValidationModel) -> ValidationResultOut:
        return ValidationResultOut(
            validation_id=str(validation.id),
            status=validation.status,
            valid_target_count=validation.valid_target_count,
            issue_count=validation.issue_count,
            issue_stats=OutboundValidationService._issue_stats(validation.issue_stats_json),
            error_message=validation.error_message,
            accepted=validation.status == "VALIDATING",
            retryable=validation.retryable,
            retry_action=(
                "RETRY_VALIDATION"
                if validation.status == "SYSTEM_ERROR" and validation.retryable
                else "REUPLOAD"
                if validation.status == "SYSTEM_ERROR"
                else None
            ),
        )

    @staticmethod
    def issue_out(row: AiCallOutboundValidationRowModel) -> ValidationIssueOut:
        duplicates = [row.duplicate_row_number] if row.duplicate_row_number else []
        return ValidationIssueOut(
            issue_id=str(row.id),
            row_number=row.row_number,
            phone_number=row.phone_number,
            customer_name=row.customer_name,
            reasons=OutboundValidationService._reasons(row.reasons_json),
            duplicate_row_numbers=duplicates,
        )

    @staticmethod
    async def _get_validation(
        db: AsyncSession,
        tenant_id: str,
        validation_id: int,
    ) -> AiCallOutboundValidationModel:
        validation = await db.scalar(
            select(AiCallOutboundValidationModel).where(
                AiCallOutboundValidationModel.tenant_id == tenant_id,
                AiCallOutboundValidationModel.id == validation_id,
            )
        )
        if validation is None:
            raise CustomException(
                msg="校验结果不存在",
                status_code=status.HTTP_404_NOT_FOUND,
            )
        return validation

    @staticmethod
    def _cell_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @classmethod
    def _value_at(cls, values: tuple[Any, ...], index: int | None) -> str:
        if index is None or index >= len(values):
            return ""
        return cls._cell_text(values[index])

    @staticmethod
    def _reasons_json(reasons: list[str]) -> str:
        return json.dumps(reasons, ensure_ascii=False, separators=(",", ":"))

    @staticmethod
    def _reasons(value: str | None) -> list[str]:
        if not value:
            return []
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return []
        return [str(item) for item in parsed] if isinstance(parsed, list) else []

    @staticmethod
    def _issue_stats(value: str | None) -> dict[str, int]:
        if not value:
            return {}
        try:
            parsed = json.loads(value)
        except json.JSONDecodeError:
            return {}
        if not isinstance(parsed, dict):
            return {}
        return {str(key): int(count) for key, count in parsed.items()}

    @staticmethod
    def _delete_temp_file(path: str | None) -> bool:
        if not path:
            return True
        candidate = Path(path)
        temp_root = Path(tempfile.gettempdir()).resolve()
        try:
            resolved = candidate.resolve()
            if resolved.parent != temp_root or not resolved.name.startswith("ai-call-outbound-"):
                log.error(f"拒绝删除非 outbound 临时文件: {resolved}")
                return False
            resolved.unlink(missing_ok=True)
            return True
        except OSError as exc:
            log.warning(f"删除 outbound 临时文件失败 path={candidate}: {exc}")
            return False
